from __future__ import annotations

import gzip
import json
import math
import os
import re
import shutil
import warnings
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from typing import Any, Iterable, Iterator, Sequence, TextIO

from .version import VERSION


COLLECTIONS = (
    "analyses",
    "biosamples",
    "cohorts",
    "datasets",
    "genomicVariations",
    "individuals",
    "runs",
)
METADATA_COLLECTIONS = tuple(
    collection for collection in COLLECTIONS if collection != "genomicVariations"
)
NUMBER_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$")
ARRAY_HEADER_RE = re.compile(r"^[A-Za-z0-9-]+_")
SCHEMA_VERSION_RE = re.compile(r"^v\d+\.\d+\.\d+$")
SCHEMA_ROOT = Path(__file__).resolve().parent / "schemas"
RESET = "\033[0m"
BOLD = "\033[1m"
BLUE = "\033[34m"
CYAN = "\033[36m"
GREEN = "\033[32m"
RED = "\033[31m"
YELLOW = "\033[33m"
COLLECTION_EMOJI = {
    "analyses": "🗂 ",
    "biosamples": "🧪 ",
    "cohorts": "👥 ",
    "datasets": "📦 ",
    "genomicVariations": "🧬 ",
    "individuals": "🧍 ",
    "runs": "🧫 ",
}


class ValidatorError(RuntimeError):
    pass


@dataclass(frozen=True)
class ValidationIssue:
    collection: str
    row: int
    message: str


@dataclass(frozen=True)
class CollectionReport:
    name: str
    checked: int
    written: Path | None = None


@dataclass(frozen=True)
class ValidationReport:
    checked: int
    written: tuple[Path, ...]
    issues: tuple[ValidationIssue, ...]
    collections: tuple[CollectionReport, ...] = ()

    @property
    def ok(self) -> bool:
        return not self.issues


def current_schema_version() -> str:
    pointer = SCHEMA_ROOT / "CURRENT"
    try:
        version = pointer.read_text(encoding="utf-8").strip()
    except OSError as exc:
        raise ValidatorError(f"Cannot read current schema version: {pointer}") from exc
    if not SCHEMA_VERSION_RE.fullmatch(version):
        raise ValidatorError(f"Invalid current schema version in {pointer}: {version!r}")
    return version


def default_schema_dir() -> Path:
    version = current_schema_version()
    directory = SCHEMA_ROOT / version
    if not directory.is_dir():
        raise ValidatorError(
            f"Current schema directory does not exist for {version}: {directory}"
        )
    return directory


def template_path() -> Path:
    return Path(__file__).resolve().parent / "resources" / "Beacon-v2-Models_template.xlsx"


def export_template(destination: Path) -> Path:
    if destination.exists():
        raise ValidatorError(f"Template destination already exists: {destination}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    source = resources.files("bff_tools").joinpath(
        "resources", "Beacon-v2-Models_template.xlsx"
    )
    with resources.as_file(source) as source_path:
        shutil.copyfile(source_path, destination)
    return destination


def _coerce_scalar(value: Any) -> Any:
    if not isinstance(value, str):
        if isinstance(value, float) and not math.isfinite(value):
            return str(value)
        return value

    lowered = value.lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    if NUMBER_RE.fullmatch(value):
        try:
            return float(value) if any(char in value for char in ".eE") else int(value)
        except ValueError:
            return value
    return value


def _decode_cell(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip().startswith(("[", "{")):
        try:
            return json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValidatorError(f"Invalid JSON cell value: {exc.msg}") from exc
    return _coerce_scalar(value)


def _has_value(value: Any) -> bool:
    return value is not None and value != ""


def _set_nested(target: dict[str, Any], path: Sequence[str], value: Any) -> None:
    if not path or any(not item for item in path):
        raise ValidatorError("Workbook header contains an empty path component")
    cursor = target
    for item in path[:-1]:
        existing = cursor.get(item)
        if existing is None:
            existing = {}
            cursor[item] = existing
        if not isinstance(existing, dict):
            raise ValidatorError(f"Workbook columns assign incompatible values to '{item}'")
        cursor = existing
    cursor[path[-1]] = value


def _cell_values(value: Any) -> list[Any]:
    decoded = _decode_cell(value)
    if isinstance(decoded, list):
        return decoded
    if isinstance(value, str):
        return [_coerce_scalar(item.strip()) for item in value.split(",")]
    return [decoded]


def row_to_document(headers: Sequence[str], values: Sequence[Any]) -> dict[str, Any]:
    document: dict[str, Any] = {}
    array_values: dict[str, dict[int, dict[str, Any]]] = {}

    for header, value in zip(headers, values):
        if not header or not _has_value(value) or ARRAY_HEADER_RE.match(header):
            continue
        _set_nested(document, header.split("."), _decode_cell(value))

    for header, value in zip(headers, values):
        if not header or not _has_value(value) or not ARRAY_HEADER_RE.match(header):
            continue
        root, remainder = header.split("_", 1)
        path = tuple(part for part in re.split(r"[._]", remainder) if part)
        if not path:
            raise ValidatorError(f"Array header requires a child field: {header}")
        for index, item in enumerate(_cell_values(value)):
            element = array_values.setdefault(root, {}).setdefault(index, {})
            _set_nested(element, path, item)

    for root, indexed_values in array_values.items():
        document[root] = [indexed_values[index] for index in sorted(indexed_values)]
    return document


def _normalise_headers(values: Sequence[Any]) -> tuple[str, ...]:
    headers = []
    for value in values:
        if value is None:
            headers.append("")
        else:
            headers.append("".join(str(value).split()))
    return tuple(headers)


def _load_schema(schema_dir: Path, collection: str) -> dict[str, Any]:
    path = schema_dir / collection / "defaultSchema.json"
    try:
        schema = json.loads(path.read_text(encoding="utf-8"))
    except OSError as exc:
        raise ValidatorError(f"Cannot read schema for {collection}: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ValidatorError(f"Invalid JSON schema {path}: {exc}") from exc
    if not isinstance(schema, dict):
        raise ValidatorError(f"Schema must contain a JSON object: {path}")
    return schema


def _schema_validator(schema: dict[str, Any]) -> Any:
    try:
        from jsonschema.validators import validator_for
    except ImportError as exc:
        raise ValidatorError(
            "JSON validation requires the 'jsonschema' package; reinstall beacon2-cbi-tools"
        ) from exc

    validator_class = validator_for(schema)
    # The dereferenced upstream genomic-variation schema contains annotation
    # keywords whose values are not accepted by strict schema self-validation.
    # They do not affect instance validation, matching the legacy validator.
    return validator_class(schema)


def _format_schema_error(error: Any) -> str:
    path = ".".join(str(part) for part in error.absolute_path)
    return f"{path or '$'}: {error.message}"


def _validate_document(
    validator: Any,
    collection: str,
    row: int,
    document: Any,
) -> list[ValidationIssue]:
    errors = sorted(
        validator.iter_errors(document),
        key=lambda error: (tuple(str(part) for part in error.absolute_path), error.message),
    )
    return [
        ValidationIssue(collection=collection, row=row, message=_format_schema_error(error))
        for error in errors
    ]


def _write_collection(path: Path, documents: Sequence[dict[str, Any]]) -> None:
    path.write_text(
        # Match JSON::XS canonical/pretty output from the legacy validator.
        json.dumps(
            documents,
            ensure_ascii=False,
            indent=3,
            sort_keys=True,
            separators=(",", " : "),
        )
        + "\n",
        encoding="utf-8",
    )


def _validate_workbook(
    workbook_path: Path,
    schema_dir: Path,
    output_dir: Path,
    collections: Sequence[str],
    *,
    ignore_validation: bool,
    verbose: bool,
) -> ValidationReport:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidatorError(
            "XLSX validation requires the 'openpyxl' package; reinstall beacon2-cbi-tools"
        ) from exc

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
            )
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise ValidatorError(f"Cannot read XLSX workbook {workbook_path}: {exc}") from exc

    missing = [collection for collection in collections if collection not in workbook.sheetnames]
    if missing:
        workbook.close()
        raise ValidatorError(f"Workbook is missing sheet(s): {', '.join(missing)}")

    checked = 0
    written: list[Path] = []
    issues: list[ValidationIssue] = []
    collection_reports: list[CollectionReport] = []
    try:
        for collection in collections:
            worksheet = workbook[collection]
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = _normalise_headers(next(rows))
            except StopIteration as exc:
                raise ValidatorError(f"Workbook sheet is empty: {collection}") from exc
            if not any(headers):
                raise ValidatorError(f"Workbook sheet has no headers: {collection}")

            documents: list[dict[str, Any]] = []
            collection_issues: list[ValidationIssue] = []
            destination: Path | None = None
            validator = _schema_validator(_load_schema(schema_dir, collection))
            for worksheet_row, values in enumerate(rows, start=2):
                if not any(_has_value(value) for value in values):
                    continue
                try:
                    document = row_to_document(headers, values)
                except ValidatorError as exc:
                    collection_issues.append(
                        ValidationIssue(collection, worksheet_row, str(exc))
                    )
                    continue
                documents.append(document)
                checked += 1
                collection_issues.extend(
                    _validate_document(validator, collection, worksheet_row, document)
                )
                if verbose and checked % 100 == 0:
                    print(f"  {checked} workbook rows checked")

            if not documents:
                collection_issues.append(
                    ValidationIssue(collection, 1, "sheet contains no data rows")
                )
            issues.extend(collection_issues)
            if not collection_issues or ignore_validation:
                destination = output_dir / f"{collection}.json"
                _write_collection(destination, documents)
                written.append(destination)
            collection_reports.append(
                CollectionReport(collection, len(documents), destination)
            )
    finally:
        workbook.close()

    return ValidationReport(
        checked,
        tuple(written),
        tuple(issues),
        tuple(collection_reports),
    )


def _open_json(path: Path) -> TextIO:
    if path.suffix == ".gz":
        return gzip.open(path, "rt", encoding="utf-8")
    return path.open("r", encoding="utf-8")


def _read_json_array(path: Path) -> list[Any]:
    try:
        with _open_json(path) as handle:
            value = json.load(handle)
    except OSError as exc:
        raise ValidatorError(f"Cannot read JSON input {path}: {exc}") from exc
    except json.JSONDecodeError as exc:
        raise ValidatorError(f"Invalid JSON input {path}: {exc}") from exc
    if not isinstance(value, list):
        raise ValidatorError(f"JSON collection must contain an array: {path}")
    return value


def _read_streamed_array(path: Path) -> Iterator[tuple[int, Any]]:
    try:
        with _open_json(path) as handle:
            if path.name.endswith((".jsonl", ".jsonl.gz")):
                row = 0
                for line_number, raw_line in enumerate(handle, start=1):
                    line = raw_line.strip()
                    if not line:
                        continue
                    row += 1
                    try:
                        yield row, json.loads(line)
                    except json.JSONDecodeError as exc:
                        raise ValidatorError(
                            f"Invalid JSON Lines record on line {line_number} of {path}: {exc}"
                        ) from exc
                return

            started = False
            row = 0
            for line_number, raw_line in enumerate(handle, start=1):
                line = raw_line.strip()
                if not line:
                    continue
                if not started:
                    if line != "[":
                        raise ValidatorError(
                            f"Expected '[' on line {line_number} of streamed JSON input"
                        )
                    started = True
                    continue
                if line == "]":
                    return
                if line.endswith(","):
                    line = line[:-1]
                row += 1
                try:
                    yield row, json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValidatorError(
                        f"Invalid streamed JSON on line {line_number} of {path}: {exc}"
                    ) from exc
    except OSError as exc:
        raise ValidatorError(f"Cannot read JSON input {path}: {exc}") from exc
    raise ValidatorError(f"Streamed JSON array is not closed: {path}")


def _collection_from_path(path: Path, include_genomic: bool, streamed: bool) -> str:
    name = path.name
    jsonl = name.endswith((".jsonl", ".jsonl.gz"))
    if jsonl and not streamed:
        raise ValidatorError("Use --gv-vcf to validate JSON Lines genomic variations")
    if name.endswith(".jsonl.gz"):
        name = name[:-9]
    elif name.endswith(".jsonl"):
        name = name[:-6]
    elif name.endswith(".json.gz"):
        name = name[:-8]
    elif name.endswith(".json"):
        name = name[:-5]
    else:
        raise ValidatorError(
            f"Input must end in .json, .json.gz, .jsonl, or .jsonl.gz: {path}"
        )
    if streamed and name == "genomicVariationsVcf":
        name = "genomicVariations"
    allowed = COLLECTIONS if include_genomic else METADATA_COLLECTIONS
    if name not in allowed:
        if name == "genomicVariations":
            raise ValidatorError("Use --gv or --gv-vcf to validate genomicVariations")
        raise ValidatorError(
            f"JSON filename must match a BFF collection: {', '.join(allowed)}"
        )
    return name


def _validate_json_files(
    paths: Sequence[Path],
    schema_dir: Path,
    *,
    include_genomic: bool,
    streamed: bool,
    verbose: bool,
) -> ValidationReport:
    checked = 0
    issues: list[ValidationIssue] = []
    seen: set[str] = set()
    collection_reports: list[CollectionReport] = []
    for path in paths:
        collection = _collection_from_path(path, include_genomic, streamed)
        if collection in seen:
            raise ValidatorError(f"Collection was provided more than once: {collection}")
        seen.add(collection)
        validator = _schema_validator(_load_schema(schema_dir, collection))
        if streamed and collection == "genomicVariations":
            documents: Iterable[tuple[int, Any]] = _read_streamed_array(path)
        else:
            documents = enumerate(_read_json_array(path), start=1)
        collection_checked = 0
        for row, document in documents:
            checked += 1
            collection_checked += 1
            issues.extend(_validate_document(validator, collection, row, document))
            if verbose and checked % 1000 == 0:
                print(f"  {checked} JSON documents checked")
        collection_reports.append(CollectionReport(collection, collection_checked))
    return ValidationReport(checked, (), tuple(issues), tuple(collection_reports))


def validate_inputs(
    inputs: Sequence[Path],
    *,
    schema_dir: Path | None = None,
    output_dir: Path = Path("."),
    include_genomic: bool = False,
    streamed_genomic: bool = False,
    ignore_validation: bool = False,
    verbose: bool = False,
) -> ValidationReport:
    if not inputs:
        raise ValidatorError("At least one input file is required")
    paths = tuple(path.resolve() for path in inputs)
    for path in paths:
        if not path.is_file():
            raise ValidatorError(f"Input file does not exist: {path}")
    selected_schema_dir = (schema_dir or default_schema_dir()).resolve()
    if not selected_schema_dir.is_dir():
        raise ValidatorError(f"Schema directory does not exist: {selected_schema_dir}")
    workbook_inputs = [path for path in paths if path.suffix.lower() == ".xlsx"]
    if workbook_inputs:
        if len(paths) != 1:
            raise ValidatorError("XLSX and JSON inputs cannot be mixed")
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise ValidatorError(f"Cannot create output directory {output_dir}: {exc}") from exc
        if not output_dir.is_dir():
            raise ValidatorError(f"Output path is not a directory: {output_dir}")
        collections = COLLECTIONS if include_genomic else METADATA_COLLECTIONS
        return _validate_workbook(
            paths[0],
            selected_schema_dir,
            output_dir.resolve(),
            collections,
            ignore_validation=ignore_validation,
            verbose=verbose,
        )
    return _validate_json_files(
        paths,
        selected_schema_dir,
        include_genomic=include_genomic or streamed_genomic,
        streamed=streamed_genomic,
        verbose=verbose,
    )


def _colorize(value: str, *codes: str, use_color: bool) -> str:
    if not use_color:
        return value
    return "".join(codes) + value + RESET


def _record_label(count: int) -> str:
    return "record" if count == 1 else "records"


def print_report(
    report: ValidationReport,
    *,
    ignore_validation: bool = False,
    no_color: bool = False,
    no_emoji: bool = False,
) -> None:
    use_color = not no_color and os.environ.get("ANSI_COLORS_DISABLED") != "1"
    header_prefix = "" if no_emoji else "🧬 "
    print(_colorize(f"{header_prefix}BFF Tools Validator  v{VERSION}", BOLD, CYAN, use_color=use_color))
    print(_colorize("Validate Beacon XLSX or JSON input and write BFF JSON collections", CYAN, use_color=use_color))

    grouped: dict[str, list[ValidationIssue]] = {}
    for issue in report.issues:
        grouped.setdefault(issue.collection, []).append(issue)

    for collection_report in report.collections:
        collection = collection_report.name
        issues = grouped.pop(collection, [])
        collection_icon = "" if no_emoji else COLLECTION_EMOJI.get(collection, "")
        print()
        print(_colorize(f"== {collection_icon}{collection} ==", BOLD, BLUE, use_color=use_color))
        if issues:
            prefix = "" if no_emoji else "✖ "
            print(
                _colorize(
                    f"{prefix}{collection}: {len(issues)} validation issue(s)",
                    BOLD,
                    RED,
                    use_color=use_color,
                )
            )
            for issue in issues:
                print(_colorize(f"  row {issue.row}: {issue.message}", RED, use_color=use_color))
            if ignore_validation:
                prefix = "" if no_emoji else "⚠ "
                print(
                    _colorize(
                        f"{prefix}Validation issues ignored for this collection",
                        BOLD,
                        YELLOW,
                        use_color=use_color,
                    )
                )
        else:
            prefix = "" if no_emoji else "✓ "
            print(
                _colorize(
                    f"{prefix}{collection}: validation passed "
                    f"({collection_report.checked:,} {_record_label(collection_report.checked)})",
                    BOLD,
                    GREEN,
                    use_color=use_color,
                )
            )
        if collection_report.written is not None:
            prefix = "" if no_emoji else "→ "
            print(
                _colorize(
                    f"{prefix}Wrote {collection_report.written}",
                    GREEN,
                    use_color=use_color,
                )
            )

    for collection, issues in grouped.items():
        print()
        print(_colorize(f"== {collection} ==", BOLD, BLUE, use_color=use_color))
        for issue in issues:
            print(_colorize(f"  row {issue.row}: {issue.message}", RED, use_color=use_color))

    print()
    if report.issues and ignore_validation:
        prefix = "" if no_emoji else "⚠ "
        print(
            _colorize(
                f"{prefix}Validation issues ignored; checked "
                f"{report.checked:,} {_record_label(report.checked)}",
                BOLD,
                YELLOW,
                use_color=use_color,
            )
        )
    elif report.issues:
        prefix = "" if no_emoji else "✖ "
        print(
            _colorize(
                f"{prefix}Validation failed; checked "
                f"{report.checked:,} {_record_label(report.checked)}",
                BOLD,
                RED,
                use_color=use_color,
            )
        )
    elif not report.issues:
        prefix = "" if no_emoji else "✓ "
        print(
            _colorize(
                f"{prefix}Validation passed; checked "
                f"{report.checked:,} {_record_label(report.checked)}",
                BOLD,
                GREEN,
                use_color=use_color,
            )
        )
