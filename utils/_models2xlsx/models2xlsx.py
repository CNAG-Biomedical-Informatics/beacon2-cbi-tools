#!/usr/bin/env python3
"""Generate the Beacon v2 metadata workbook from dereferenced model schemas."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Iterable, Sequence


COLLECTIONS = (
    "analyses",
    "biosamples",
    "cohorts",
    "datasets",
    "genomicVariations",
    "individuals",
    "runs",
)


def _properties(schema: dict[str, Any], *, array_items: bool = False) -> dict[str, Any]:
    if array_items:
        items = schema.get("items", {})
        return items.get("properties", {}) if isinstance(items, dict) else {}
    properties = schema.get("properties", {})
    return properties if isinstance(properties, dict) else {}


def _children(
    depth: int, property_type: object, schema: dict[str, Any]
) -> tuple[dict[str, Any], str]:
    """Mirror the historical template layout without expanding nested arrays."""
    if depth == 1 and property_type == "array":
        return _properties(schema, array_items=True), "_"
    if property_type == "array":
        return {}, "."
    return _properties(schema), "."


def _remove_parent_headers(headers: Iterable[str]) -> list[str]:
    terms = list(headers)
    return [
        term
        for term in terms
        if not any(
            other != term
            and (other.startswith(f"{term}.") or other.startswith(f"{term}_"))
            for other in terms
        )
    ]


def headers_from_schema(schema_path: Path) -> list[str]:
    """Return workbook headers using the layout of the original Perl utility."""
    with schema_path.open(encoding="utf-8") as handle:
        schema = json.load(handle)

    headers: list[str] = []
    for name_1d, value_1d in sorted(_properties(schema).items()):
        type_1d = "string" if name_1d == "variation" else value_1d.get("type")
        if type_1d not in {"array", "object"}:
            headers.append(name_1d)
            continue

        properties_2d, separator_2d = _children(1, type_1d, value_1d)
        for name_2d, value_2d in sorted(properties_2d.items()):
            header_2d = f"{name_1d}{separator_2d}{name_2d}"
            headers.append(header_2d)

            properties_3d, separator_3d = _children(
                2, value_2d.get("type"), value_2d
            )
            for name_3d, value_3d in sorted(properties_3d.items()):
                header_3d = f"{header_2d}{separator_3d}{name_3d}"
                headers.append(header_3d)

                properties_4d, separator_4d = _children(
                    3, value_3d.get("type"), value_3d
                )
                for name_4d in sorted(properties_4d):
                    headers.append(f"{header_3d}{separator_4d}{name_4d}")

    headers = sorted(_remove_parent_headers(headers))
    if schema_path.parent.name in {"cohorts", "datasets"}:
        headers.extend(("ids.individualIds", "ids.biosampleIds"))
    return headers


def collection_headers(schema_dir: Path) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for collection in COLLECTIONS:
        schema_path = schema_dir / collection / "defaultSchema.json"
        if not schema_path.is_file():
            raise FileNotFoundError(f"missing dereferenced schema: {schema_path}")
        result[collection] = headers_from_schema(schema_path)
    return result


def write_csv_files(headers: dict[str, list[str]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for collection, fields in headers.items():
        with (output_dir / f"{collection}.csv").open(
            "w", encoding="utf-8", newline=""
        ) as handle:
            csv.writer(handle, lineterminator="\n").writerow(fields)


def write_workbook(headers: dict[str, list[str]], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - normal project installs include it
        raise RuntimeError(
            "workbook generation requires openpyxl; install beacon2-cbi-tools first"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for collection, fields in headers.items():
        worksheet = workbook.create_sheet(collection)
        worksheet.append(fields)
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for index, field in enumerate(fields, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(field) + 2, 12), 60
            )
    workbook.save(output_path)


def default_schema_dir() -> Path:
    schema_root = Path(__file__).resolve().parents[2] / "src" / "bff_tools" / "schemas"
    pointer = schema_root / "CURRENT"
    try:
        version = pointer.read_text(encoding="utf-8").strip()
    except OSError as exc:
        raise FileNotFoundError(f"cannot read current schema version: {pointer}") from exc
    schema_dir = schema_root / version
    if not schema_dir.is_dir():
        raise FileNotFoundError(
            f"current schema directory does not exist for {version}: {schema_dir}"
        )
    return schema_dir


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the Beacon-v2-Models XLSX template from the packaged "
            "dereferenced schemas."
        )
    )
    parser.add_argument(
        "--schema-dir",
        type=Path,
        default=default_schema_dir(),
        help="directory containing one defaultSchema.json per Beacon collection",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).with_name("Beacon-v2-Models_template.xlsx"),
        help="generated XLSX path",
    )
    parser.add_argument(
        "--csv-dir",
        type=Path,
        default=Path(__file__).with_name("data"),
        help="directory for the intermediate one-row CSV files",
    )
    parser.add_argument(
        "--no-csv",
        action="store_true",
        help="do not retain the intermediate CSV files",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    headers = collection_headers(args.schema_dir)
    if not args.no_csv:
        write_csv_files(headers, args.csv_dir)
    write_workbook(headers, args.output)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
