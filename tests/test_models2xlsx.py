from __future__ import annotations

import importlib.util
import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
UTILITY_PATH = ROOT / "utils" / "_models2xlsx" / "models2xlsx.py"
LEGACY_PARSER = (
    ROOT / "utils" / "_models2xlsx" / "test" / "parse_defaultSchema.pl"
)
SPEC = importlib.util.spec_from_file_location("models2xlsx", UTILITY_PATH)
if SPEC is None or SPEC.loader is None:  # pragma: no cover - fixed repository path
    raise RuntimeError(f"cannot load {UTILITY_PATH}")
MODELS2XLSX = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODELS2XLSX)
SCHEMA_DIR = MODELS2XLSX.default_schema_dir()


class Models2XlsxTests(unittest.TestCase):
    def test_headers_preserve_legacy_template_structure(self) -> None:
        headers = MODELS2XLSX.collection_headers(SCHEMA_DIR)
        self.assertEqual(
            {collection: len(fields) for collection, fields in headers.items()},
            {
                "analyses": 9,
                "biosamples": 59,
                "cohorts": 52,
                "datasets": 10,
                "genomicVariations": 28,
                "individuals": 81,
                "runs": 12,
            },
        )
        self.assertEqual(
            headers["cohorts"][-2:], ["ids.individualIds", "ids.biosampleIds"]
        )

    def test_generator_writes_workbook_and_csv_headers(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "template.xlsx"
            csv_dir = Path(tmpdir) / "csv"
            result = MODELS2XLSX.main(
                [
                    "--schema-dir",
                    str(SCHEMA_DIR),
                    "--output",
                    str(output),
                    "--csv-dir",
                    str(csv_dir),
                ]
            )

            self.assertEqual(result, 0)
            self.assertTrue(output.is_file())
            self.assertEqual(
                load_workbook(output, read_only=True).sheetnames,
                list(MODELS2XLSX.COLLECTIONS),
            )
            self.assertEqual(
                (csv_dir / "analyses.csv").read_text(encoding="utf-8").strip(),
                "aligner,analysisDate,biosampleId,id,individualId,pipelineName,"
                "pipelineRef,runId,variantCaller",
            )

    def test_headers_match_legacy_perl_parser_when_available(self) -> None:
        perl = shutil.which("perl")
        if perl is None:
            self.skipTest("Perl is not installed")
        dependency_check = subprocess.run(
            [perl, "-MJSON::XS", "-MPath::Tiny", "-e", "1"],
            capture_output=True,
            check=False,
            text=True,
        )
        if dependency_check.returncode:
            self.skipTest("legacy Perl parser dependencies are not installed")

        for collection in MODELS2XLSX.COLLECTIONS:
            schema_path = SCHEMA_DIR / collection / "defaultSchema.json"
            completed = subprocess.run(
                [perl, str(LEGACY_PARSER), str(schema_path)],
                capture_output=True,
                check=True,
                text=True,
            )
            self.assertEqual(
                MODELS2XLSX.headers_from_schema(schema_path),
                completed.stdout.strip().split(","),
                collection,
            )


if __name__ == "__main__":
    unittest.main()
