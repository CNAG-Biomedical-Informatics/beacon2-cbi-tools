from __future__ import annotations

import contextlib
import builtins
import gzip
import io
import json
import tempfile
import unittest
import warnings
from pathlib import Path
from unittest import mock

import sys

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

CINECA_ROOT = ROOT / "CINECA_synthetic_cohort_EUROPE_UK1"
CINECA_CURRENT = CINECA_ROOT / "current"
CINECA_V2 = CINECA_ROOT / "versions" / "v2.0.0"
SCHEMA_V2 = SRC / "bff_tools" / "schemas" / "v2.0.0"

import bff_tools.validator as validator  # noqa: E402
from bff_tools.validator import (  # noqa: E402
    CollectionReport,
    ValidationIssue,
    ValidationReport,
    ValidatorError,
    export_template,
    print_report,
    row_to_document,
    validate_inputs,
)


class ValidatorTests(unittest.TestCase):
    @staticmethod
    def _write_schema(schema_dir: Path, collection: str = "individuals") -> None:
        target = schema_dir / collection
        target.mkdir(parents=True, exist_ok=True)
        (target / "defaultSchema.json").write_text(
            json.dumps(
                {
                    "type": "object",
                    "required": ["id"],
                    "properties": {"id": {"type": "string"}},
                }
            ),
            encoding="utf-8",
        )

    def test_row_to_document_preserves_workbook_mapping(self) -> None:
        headers = (
            "id",
            "sex.id",
            "active",
            "count",
            "tags",
            "measures_assayCode.id",
            "measures_assayCode.label",
        )
        values = (
            "person-1",
            "NCIT:C20197",
            "true",
            "4",
            '["alpha", "beta"]',
            "LOINC:1,LOINC:2",
            "First,Second",
        )
        self.assertEqual(
            row_to_document(headers, values),
            {
                "id": "person-1",
                "sex": {"id": "NCIT:C20197"},
                "active": True,
                "count": 4,
                "tags": ["alpha", "beta"],
                "measures": [
                    {"assayCode": {"id": "LOINC:1", "label": "First"}},
                    {"assayCode": {"id": "LOINC:2", "label": "Second"}},
                ],
            },
        )

    def test_row_to_document_rejects_malformed_json_cell(self) -> None:
        with self.assertRaisesRegex(ValidatorError, "Invalid JSON cell value"):
            row_to_document(("id", "values"), ("person-1", "[broken"))

    def test_json_validation_reports_schema_errors(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "individuals.json"
            path.write_text('[{"id": "person-1"}]\n', encoding="utf-8")
            report = validate_inputs([path])
        self.assertFalse(report.ok)
        self.assertTrue(any("sex" in issue.message for issue in report.issues))

    def test_json_validation_accepts_valid_collection(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "individuals.json"
            path.write_text(
                json.dumps(
                    [
                        {
                            "id": "person-1",
                            "sex": {"id": "NCIT:C20197", "label": "male"},
                        }
                    ]
                ),
                encoding="utf-8",
            )
            report = validate_inputs([path])
        self.assertTrue(report.ok)
        self.assertEqual(report.checked, 1)

    def test_report_restores_collection_status_and_respects_output_flags(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "individuals.json"
            path.write_text(
                '[{"id":"person-1","sex":{"id":"NCIT:C20197","label":"male"}}]\n',
                encoding="utf-8",
            )
            report = validate_inputs([path])

        output = io.StringIO()
        with contextlib.redirect_stdout(output):
            print_report(report, no_color=True)
        rendered = output.getvalue()
        self.assertIn("🧬 BFF Tools Validator", rendered)
        self.assertIn("== 🧍 individuals ==", rendered)
        self.assertIn("✓ individuals: validation passed (1 record)", rendered)
        self.assertNotIn("\033[", rendered)

        plain_output = io.StringIO()
        with contextlib.redirect_stdout(plain_output):
            print_report(report, no_color=True, no_emoji=True)
        self.assertNotIn("🧬", plain_output.getvalue())
        self.assertNotIn("✓", plain_output.getvalue())

    def test_mixed_xlsx_and_json_inputs_are_rejected(self) -> None:
        workbook = CINECA_CURRENT / "Beacon-v2-Models_CINECA_UK1.xlsx"
        individuals = CINECA_CURRENT / "bff" / "individuals.json"
        with self.assertRaisesRegex(ValidatorError, "cannot be mixed"):
            validate_inputs([workbook, individuals])

    def test_cineca_workbook_matches_perl_generated_bff_byte_for_byte(self) -> None:
        workbook = CINECA_V2 / "Beacon-v2-Models_CINECA_UK1.xlsx"
        expected_dir = CINECA_V2 / "bff"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Unknown extension is not supported")
                report = validate_inputs(
                    [workbook], schema_dir=SCHEMA_V2, output_dir=output_dir
                )
            self.assertTrue(report.ok)
            self.assertEqual(report.checked, 10018)
            for collection in (
                "analyses",
                "biosamples",
                "cohorts",
                "datasets",
                "individuals",
                "runs",
            ):
                actual = (output_dir / f"{collection}.json").read_bytes()
                expected = (expected_dir / f"{collection}.json").read_bytes()
                self.assertEqual(actual, expected, collection)

    def test_workbook_output_directory_is_created(self) -> None:
        workbook = CINECA_CURRENT / "Beacon-v2-Models_CINECA_UK1.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "nested" / "bff"
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Unknown extension is not supported")
                report = validate_inputs([workbook], output_dir=output_dir)
            self.assertTrue(report.ok)
            self.assertTrue((output_dir / "individuals.json").is_file())

    def test_template_export_writes_packaged_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            destination = Path(tmpdir) / "metadata.xlsx"
            export_template(destination)
            self.assertTrue(destination.is_file())
            self.assertGreater(destination.stat().st_size, 1000)
            with self.assertRaisesRegex(ValidatorError, "already exists"):
                export_template(destination)

    def test_scalar_header_and_nested_mapping_edge_cases(self) -> None:
        self.assertTrue(validator.template_path().is_file())
        self.assertEqual(validator._coerce_scalar(float("inf")), "inf")
        self.assertIs(validator._coerce_scalar("false"), False)
        self.assertEqual(validator._coerce_scalar("1.5e2"), 150.0)
        self.assertEqual(validator._normalise_headers((None, " sex . id ")), ("", "sex.id"))

        with self.assertRaisesRegex(ValidatorError, "empty path component"):
            validator._set_nested({}, (), "value")
        with self.assertRaisesRegex(ValidatorError, "incompatible values"):
            row_to_document(("sex", "sex.id"), ("male", "NCIT:C20197"))
        with self.assertRaisesRegex(ValidatorError, "requires a child field"):
            row_to_document(("measures_",), ("value",))

    def test_schema_loading_reports_missing_invalid_and_non_object_schemas(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            with self.assertRaisesRegex(ValidatorError, "Cannot read schema"):
                validator._load_schema(root, "individuals")

            schema_dir = root / "individuals"
            schema_dir.mkdir()
            schema_path = schema_dir / "defaultSchema.json"
            schema_path.write_text("{broken", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "Invalid JSON schema"):
                validator._load_schema(root, "individuals")

            schema_path.write_text("[]", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "JSON object"):
                validator._load_schema(root, "individuals")

        original_import = builtins.__import__

        def blocked_import(name, *args, **kwargs):
            if name == "jsonschema.validators":
                raise ImportError("blocked for test")
            return original_import(name, *args, **kwargs)

        with mock.patch("builtins.__import__", side_effect=blocked_import):
            with self.assertRaisesRegex(ValidatorError, "jsonschema"):
                validator._schema_validator({"type": "object"})

    def test_json_readers_support_gzip_and_report_invalid_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            compressed = tmp / "individuals.json.gz"
            with gzip.open(compressed, "wt", encoding="utf-8") as handle:
                json.dump([{"id": "one"}], handle)
            self.assertEqual(validator._read_json_array(compressed), [{"id": "one"}])

            malformed = tmp / "malformed.json"
            malformed.write_text("[broken", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "Invalid JSON input"):
                validator._read_json_array(malformed)
            non_array = tmp / "non-array.json"
            non_array.write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "must contain an array"):
                validator._read_json_array(non_array)
            with self.assertRaisesRegex(ValidatorError, "Cannot read JSON input"):
                validator._read_json_array(tmp / "missing.json")

    def test_streamed_json_reader_handles_records_and_failure_modes(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            valid = tmp / "valid.json"
            valid.write_text('\n[\n{"id":"one"},\n\n{"id":"two"}\n]\n', encoding="utf-8")
            self.assertEqual(
                list(validator._read_streamed_array(valid)),
                [(1, {"id": "one"}), (2, {"id": "two"})],
            )

            jsonl = tmp / "valid.jsonl"
            jsonl.write_text('{"id":"one"}\n{"id":"two"}\n', encoding="utf-8")
            self.assertEqual(
                list(validator._read_streamed_array(jsonl)),
                [(1, {"id": "one"}), (2, {"id": "two"})],
            )

            cases = {
                "bad-start.json": '{}\n',
                "bad-record.json": '[\n{"id":}\n]\n',
                "unclosed.json": '[\n{"id":"one"}\n',
            }
            for name, content in cases.items():
                path = tmp / name
                path.write_text(content, encoding="utf-8")
                with self.subTest(name=name), self.assertRaises(ValidatorError):
                    list(validator._read_streamed_array(path))
            with self.assertRaisesRegex(ValidatorError, "Cannot read JSON input"):
                list(validator._read_streamed_array(tmp / "missing.json"))

    def test_collection_filename_rules_cover_aliases_and_actionable_errors(self) -> None:
        self.assertEqual(
            validator._collection_from_path(
                Path("genomicVariationsVcf.json.gz"), True, True
            ),
            "genomicVariations",
        )
        self.assertEqual(
            validator._collection_from_path(
                Path("genomicVariationsVcf.jsonl.gz"), True, True
            ),
            "genomicVariations",
        )
        with self.assertRaisesRegex(ValidatorError, "must end"):
            validator._collection_from_path(Path("individuals.txt"), False, False)
        with self.assertRaisesRegex(ValidatorError, "Use --gv-vcf"):
            validator._collection_from_path(
                Path("genomicVariationsVcf.jsonl"), True, False
            )
        with self.assertRaisesRegex(ValidatorError, "Use --gv"):
            validator._collection_from_path(Path("genomicVariations.json"), False, False)
        with self.assertRaisesRegex(ValidatorError, "filename must match"):
            validator._collection_from_path(Path("unknown.json"), False, False)

    def test_json_validation_covers_streaming_duplicates_and_progress(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            schema_dir = tmp / "schemas"
            self._write_schema(schema_dir)
            self._write_schema(schema_dir, "genomicVariations")

            first_dir = tmp / "first"
            second_dir = tmp / "second"
            first_dir.mkdir()
            second_dir.mkdir()
            first = first_dir / "individuals.json"
            second = second_dir / "individuals.json"
            first.write_text('[{"id":"one"}]', encoding="utf-8")
            second.write_text('[{"id":"two"}]', encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "more than once"):
                validate_inputs([first, second], schema_dir=schema_dir)

            many = tmp / "individuals.json"
            many.write_text(
                json.dumps([{"id": str(index)} for index in range(1000)]),
                encoding="utf-8",
            )
            output = io.StringIO()
            with contextlib.redirect_stdout(output):
                report = validate_inputs([many], schema_dir=schema_dir, verbose=True)
            self.assertEqual(report.checked, 1000)
            self.assertIn("1000 JSON documents checked", output.getvalue())

            streamed = tmp / "genomicVariationsVcf.json"
            streamed.write_text('[\n{"id":"variant-1"}\n]\n', encoding="utf-8")
            report = validate_inputs(
                [streamed], schema_dir=schema_dir, streamed_genomic=True
            )
            self.assertEqual(report.checked, 1)

    def test_validate_inputs_reports_missing_paths_and_invalid_output_directory(self) -> None:
        with self.assertRaisesRegex(ValidatorError, "At least one"):
            validate_inputs([])
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            with self.assertRaisesRegex(ValidatorError, "does not exist"):
                validate_inputs([tmp / "missing.json"])

            data = tmp / "individuals.json"
            data.write_text("[]", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "Schema directory"):
                validate_inputs([data], schema_dir=tmp / "missing-schemas")

            workbook = CINECA_CURRENT / "Beacon-v2-Models_CINECA_UK1.xlsx"
            output_file = tmp / "output-file"
            output_file.write_text("not a directory", encoding="utf-8")
            with mock.patch.object(Path, "mkdir", return_value=None):
                with self.assertRaisesRegex(ValidatorError, "not a directory"):
                    validate_inputs([workbook], output_dir=output_file)

            blocked_output = tmp / "parent-file" / "output"
            (tmp / "parent-file").write_text("blocked", encoding="utf-8")
            with self.assertRaisesRegex(ValidatorError, "Cannot create output"):
                validate_inputs([workbook], output_dir=blocked_output)

    def test_workbook_validation_reports_sheet_and_mapping_problems(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            schema_dir = tmp / "schemas"
            self._write_schema(schema_dir)
            output_dir = tmp / "output"
            output_dir.mkdir()

            missing_sheet = tmp / "missing-sheet.xlsx"
            Workbook().save(missing_sheet)
            with self.assertRaisesRegex(ValidatorError, "missing sheet"):
                validator._validate_workbook(
                    missing_sheet,
                    schema_dir,
                    output_dir,
                    ("individuals",),
                    ignore_validation=False,
                    verbose=False,
                )

            empty = Workbook()
            empty.active.title = "individuals"
            empty_path = tmp / "empty.xlsx"
            empty.save(empty_path)
            with self.assertRaisesRegex(ValidatorError, "sheet is empty"):
                validator._validate_workbook(
                    empty_path,
                    schema_dir,
                    output_dir,
                    ("individuals",),
                    ignore_validation=False,
                    verbose=False,
                )

            no_headers = Workbook()
            no_headers.active.title = "individuals"
            no_headers.active.append(["   "])
            no_headers_path = tmp / "no-headers.xlsx"
            no_headers.save(no_headers_path)
            with self.assertRaisesRegex(ValidatorError, "no headers"):
                validator._validate_workbook(
                    no_headers_path,
                    schema_dir,
                    output_dir,
                    ("individuals",),
                    ignore_validation=False,
                    verbose=False,
                )

            invalid = Workbook()
            invalid.active.title = "individuals"
            invalid.active.append(["sex", "sex.id"])
            invalid.active.append(["", ""])
            invalid.active.append(["male", "NCIT:C20197"])
            invalid_path = tmp / "invalid-mapping.xlsx"
            invalid.save(invalid_path)
            report = validator._validate_workbook(
                invalid_path,
                schema_dir,
                output_dir,
                ("individuals",),
                ignore_validation=True,
                verbose=False,
            )
            self.assertFalse(report.ok)
            self.assertEqual(report.checked, 0)
            self.assertEqual(json.loads((output_dir / "individuals.json").read_text()), [])

    def test_workbook_progress_and_optional_dependency_error(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            schema_dir = tmp / "schemas"
            self._write_schema(schema_dir)
            output_dir = tmp / "output"
            output_dir.mkdir()
            workbook = Workbook()
            workbook.active.title = "individuals"
            workbook.active.append(["id"])
            for index in range(100):
                workbook.active.append([f"person-{index}"])
            workbook_path = tmp / "many.xlsx"
            workbook.save(workbook_path)
            output = io.StringIO()
            with contextlib.redirect_stdout(output):
                report = validator._validate_workbook(
                    workbook_path,
                    schema_dir,
                    output_dir,
                    ("individuals",),
                    ignore_validation=False,
                    verbose=True,
                )
            self.assertTrue(report.ok)
            self.assertIn("100 workbook rows checked", output.getvalue())

            original_import = builtins.__import__

            def blocked_import(name, *args, **kwargs):
                if name == "openpyxl":
                    raise ImportError("blocked for test")
                return original_import(name, *args, **kwargs)

            with mock.patch("builtins.__import__", side_effect=blocked_import):
                with self.assertRaisesRegex(ValidatorError, "openpyxl"):
                    validator._validate_workbook(
                        workbook_path,
                        schema_dir,
                        output_dir,
                        ("individuals",),
                        ignore_validation=False,
                        verbose=False,
                    )

    def test_report_renders_failure_ignored_written_and_orphan_issues(self) -> None:
        report = ValidationReport(
            checked=2,
            written=(Path("individuals.json"),),
            issues=(
                ValidationIssue("individuals", 2, "sex is required"),
                ValidationIssue("unknown", 1, "orphan issue"),
            ),
            collections=(
                CollectionReport("individuals", 1, Path("individuals.json")),
            ),
        )
        ignored = io.StringIO()
        with contextlib.redirect_stdout(ignored):
            print_report(report, ignore_validation=True, no_color=False, no_emoji=False)
        rendered = ignored.getvalue()
        self.assertIn("validation issue", rendered)
        self.assertIn("Validation issues ignored", rendered)
        self.assertIn("Wrote individuals.json", rendered)
        self.assertIn("== unknown ==", rendered)
        self.assertIn("\033[", rendered)

        failed = io.StringIO()
        with contextlib.redirect_stdout(failed):
            print_report(report, no_color=True, no_emoji=True)
        self.assertIn("Validation failed; checked 2 records", failed.getvalue())


if __name__ == "__main__":
    unittest.main()
